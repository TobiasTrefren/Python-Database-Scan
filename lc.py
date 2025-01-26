from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoSuchElementException
import TKinterModernThemes as TKMT
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tkinter import *
import tkinter as tk
import time
import csv
import sys
main = Tk()

wb = load_workbook(filename='list.xlsx')
ws = wb.active
sheet2 = wb.create_sheet('Output')

driver = webdriver.Chrome()
driver.get("https://appsdoc.wi.gov/lop/home/home")


Busy = False
currow = 1
my_red = '00FF0000'
my_green = '0000FF00'

def save():
    wb.save('list.xlsx')
    print('Saved Workbook')
    
def write(fn, mn, ln, url, outcome):
    global currow
    count = 0
    if(outcome=='nodata'):
        while count <= 3:
            count += 1
            if(count==1):
                mycell = sheet2.cell(row=currow, column=count, value=fn)
                mycell.fill = PatternFill("solid", fgColor=my_green)
            elif(count==2):
                mycell = sheet2.cell(row=currow, column=count, value=mn)
                mycell.fill = PatternFill("solid", fgColor=my_green)
            elif(count==3):
                mycell = sheet2.cell(row=currow, column=count, value=ln)
                mycell.fill = PatternFill("solid", fgColor=my_green)
                save()
                currow +=1
    elif(outcome=='data'):
        while count <= 4:
            count += 1
            if(count==1):
                mycell = sheet2.cell(row=currow, column=count, value=fn)
                mycell.fill = PatternFill("solid", fgColor=my_red)
            elif(count==2):
                mycell = sheet2.cell(row=currow, column=count, value=mn)
                mycell.fill = PatternFill("solid", fgColor=my_red)
            elif(count==3):
                mycell = sheet2.cell(row=currow, column=count, value=ln)
                mycell.fill = PatternFill("solid", fgColor=my_red)
            elif(count==4):
                mycell = sheet2.cell(row=currow, column=count, value=url)
                save()
                currow += 1
          
          
def reset():
    try:
        goback = driver.find_element(By.LINK_TEXT, "New Search")
    except NoSuchElementException:
        return False      
    else:
        goback.click()
        return True
    

def search(fn, mn, ln):
    val = reset()
    while(val):
        print('Waiting')
        val = reset()
        time.sleep(.5)
    firstname = driver.find_element(By.ID, 'FIRST_NAM')
    middlename = driver.find_element(By.ID, 'MID_NAM')
    lastname = driver.find_element(By.ID, 'LAST_NAM')
    firstname.clear()
    middlename.clear()
    lastname.clear()
    firstname.send_keys(fn)
    
    if not mn:
        print('do nun')
    else:
        middlename.send_keys(mn)
    
    lastname.send_keys(ln)
    search = driver.find_element(By.LINK_TEXT, "Perform Search")
    search.click()
    try:
        em = driver.find_element(By.CLASS_NAME, "errormessage")
    except NoSuchElementException:
        url = driver.current_url
        write(fn, mn, ln, url, 'data')
    else:
        write(fn, mn, ln, 'none', 'nodata')
        

def itirateList():
    for i in range(1, ws.max_row+1):
        firstname = ws.cell(row=i, column=1).value
        middlename = ws.cell(row=i, column=2).value
        lastname = ws.cell(row=i, column=3).value
        search(firstname, middlename, lastname)
        
def leftKey(event):
         itirateList()

button = tk.Button(main, text="Start", command=itirateList, activebackground='orange', activeforeground='white', anchor="center", bd=3, bg="lightgray", disabledforeground="gray", fg="black", font=("Arial", 12),height=2, highlightbackground='black', highlightcolor='green', highlightthickness=2, justify='center', overrelief='raised', padx=10, pady=5, width=15, wraplength=100)
button.pack(padx=20, pady=20)
main.bind('<Left>', leftKey)
main.mainloop()    
    